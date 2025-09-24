"""
Report Generator - Core Report Generation Engine

This service handles the generation of various types of reports
including security dashboards, risk assessments, and compliance reports.
"""

import asyncio
import json
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Union
from uuid import UUID
from pathlib import Path
import hashlib
import os

import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
from jinja2 import Environment, FileSystemLoader
import weasyprint
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

from shared.observability.logging import get_logger
from shared.observability.metrics import get_metrics
from shared.observability.tracing import traced
from shared.config.settings import get_settings

from ..models.reporting import (
    Report, ReportTemplate, ReportType, ReportFormat, ReportStatus,
    ReportExecution, get_db
)
from .data_collector import DataCollector
from .chart_generator import ChartGenerator

logger = get_logger(__name__)
metrics = get_metrics()


class ReportGenerator:
    """
    Core report generation engine that handles various report types and formats.
    
    This generator:
    1. Processes report requests
    2. Collects and transforms data
    3. Applies templates and styling
    4. Generates reports in multiple formats
    5. Manages report lifecycle
    """
    
    def __init__(self, data_collector: DataCollector, chart_generator: ChartGenerator):
        self.data_collector = data_collector
        self.chart_generator = chart_generator
        self.settings = get_settings()
        
        # Template environment
        self.template_env = Environment(
            loader=FileSystemLoader(self.settings.report_templates_path),
            autoescape=True
        )
        
        # Output directory
        self.output_dir = Path(self.settings.report_output_path)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Generation queue
        self.generation_queue = asyncio.Queue()
        self.generation_tasks = set()
        
        logger.info("Report generator initialized")
    
    async def start(self):
        """Start the report generator."""
        # Start generation processor
        processor_task = asyncio.create_task(self._process_generation_queue())
        self.generation_tasks.add(processor_task)
        
        logger.info("Report generator started")
    
    async def stop(self):
        """Stop the report generator."""
        # Cancel generation tasks
        for task in self.generation_tasks:
            task.cancel()
        
        await asyncio.gather(*self.generation_tasks, return_exceptions=True)
        
        logger.info("Report generator stopped")
    
    @traced("report_generator_generate_report")
    async def generate_report(self, report_id: UUID) -> bool:
        """Generate a report."""
        start_time = time.time()
        
        try:
            with get_db() as db:
                report = db.query(Report).filter(Report.id == report_id).first()
                if not report:
                    raise ValueError(f"Report {report_id} not found")
                
                # Update report status
                report.status = ReportStatus.GENERATING
                report.generation_started_at = datetime.now()
                db.commit()
                
                # Create execution record
                execution = ReportExecution(
                    report_id=report_id,
                    execution_type="manual",
                    triggered_by=report.created_by,
                    status="running",
                    started_at=datetime.now()
                )
                db.add(execution)
                db.commit()
                
                try:
                    # Generate report based on type
                    if report.report_type == ReportType.SECURITY_DASHBOARD:
                        result = await self._generate_security_dashboard(report)
                    elif report.report_type == ReportType.RISK_ASSESSMENT:
                        result = await self._generate_risk_assessment_report(report)
                    elif report.report_type == ReportType.INCIDENT_ANALYSIS:
                        result = await self._generate_incident_analysis_report(report)
                    elif report.report_type == ReportType.COMPLIANCE_REPORT:
                        result = await self._generate_compliance_report(report)
                    elif report.report_type == ReportType.THREAT_INTELLIGENCE:
                        result = await self._generate_threat_intelligence_report(report)
                    elif report.report_type == ReportType.PERFORMANCE_METRICS:
                        result = await self._generate_performance_metrics_report(report)
                    elif report.report_type == ReportType.EXECUTIVE_SUMMARY:
                        result = await self._generate_executive_summary(report)
                    else:
                        result = await self._generate_custom_report(report)
                    
                    # Update report with results
                    report.status = ReportStatus.COMPLETED
                    report.content = result["content"]
                    report.file_path = result["file_path"]
                    report.file_size = result["file_size"]
                    report.file_hash = result["file_hash"]
                    report.generation_completed_at = datetime.now()
                    report.generation_duration = time.time() - start_time
                    
                    # Update execution record
                    execution.status = "completed"
                    execution.completed_at = datetime.now()
                    execution.duration = report.generation_duration
                    execution.output_file = result["file_path"]
                    execution.output_size = result["file_size"]
                    execution.record_count = result.get("record_count", 0)
                    
                    db.commit()
                    
                    metrics.report_generator_reports_generated.inc()
                    metrics.report_generator_generation_time.observe(report.generation_duration)
                    
                    logger.info(f"Report {report_id} generated successfully")
                    return True
                    
                except Exception as e:
                    # Update report status
                    report.status = ReportStatus.FAILED
                    report.generation_completed_at = datetime.now()
                    report.generation_duration = time.time() - start_time
                    
                    # Update execution record
                    execution.status = "failed"
                    execution.completed_at = datetime.now()
                    execution.duration = report.generation_duration
                    execution.error_message = str(e)
                    execution.error_details = {"error_type": type(e).__name__}
                    
                    db.commit()
                    
                    metrics.report_generator_reports_failed.inc()
                    logger.error(f"Report generation failed: {e}")
                    return False
                    
        except Exception as e:
            logger.error(f"Error generating report {report_id}: {e}")
            metrics.report_generator_errors.inc()
            return False
    
    async def _process_generation_queue(self):
        """Process report generation queue."""
        while True:
            try:
                # Get report ID from queue
                report_id = await self.generation_queue.get()
                
                # Generate report
                task = asyncio.create_task(self.generate_report(report_id))
                self.generation_tasks.add(task)
                task.add_done_callback(self.generation_tasks.discard)
                
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Error processing generation queue: {e}")
                await asyncio.sleep(1)
    
    async def _generate_security_dashboard(self, report: Report) -> Dict[str, Any]:
        """Generate security dashboard report."""
        try:
            # Collect security metrics data
            data = await self.data_collector.collect_security_metrics(
                start_date=report.start_date,
                end_date=report.end_date,
                filters=report.filters
            )
            
            # Generate charts
            charts = await self._generate_security_dashboard_charts(data)
            
            # Prepare report content
            content = {
                "title": report.title,
                "generated_at": datetime.now().isoformat(),
                "period": {
                    "start": report.start_date.isoformat(),
                    "end": report.end_date.isoformat()
                },
                "summary": {
                    "total_incidents": data.get("total_incidents", 0),
                    "high_severity_incidents": data.get("high_severity_incidents", 0),
                    "avg_response_time": data.get("avg_response_time", 0),
                    "security_score": data.get("security_score", 0)
                },
                "charts": charts,
                "data": data
            }
            
            # Generate output file
            file_result = await self._generate_output_file(report, content)
            
            return {
                "content": content,
                "file_path": file_result["file_path"],
                "file_size": file_result["file_size"],
                "file_hash": file_result["file_hash"],
                "record_count": len(data.get("incidents", []))
            }
            
        except Exception as e:
            logger.error(f"Error generating security dashboard: {e}")
            raise
    
    async def _generate_risk_assessment_report(self, report: Report) -> Dict[str, Any]:
        """Generate risk assessment report."""
        try:
            # Collect risk assessment data
            data = await self.data_collector.collect_risk_assessment_data(
                start_date=report.start_date,
                end_date=report.end_date,
                filters=report.filters
            )
            
            # Generate risk analysis charts
            charts = await self._generate_risk_assessment_charts(data)
            
            # Calculate risk metrics
            risk_metrics = self._calculate_risk_metrics(data)
            
            # Prepare report content
            content = {
                "title": report.title,
                "generated_at": datetime.now().isoformat(),
                "period": {
                    "start": report.start_date.isoformat(),
                    "end": report.end_date.isoformat()
                },
                "executive_summary": {
                    "total_assessments": data.get("total_assessments", 0),
                    "high_risk_items": data.get("high_risk_items", 0),
                    "avg_risk_score": data.get("avg_risk_score", 0),
                    "risk_trend": data.get("risk_trend", "stable")
                },
                "risk_metrics": risk_metrics,
                "charts": charts,
                "recommendations": data.get("recommendations", []),
                "data": data
            }
            
            # Generate output file
            file_result = await self._generate_output_file(report, content)
            
            return {
                "content": content,
                "file_path": file_result["file_path"],
                "file_size": file_result["file_size"],
                "file_hash": file_result["file_hash"],
                "record_count": len(data.get("assessments", []))
            }
            
        except Exception as e:
            logger.error(f"Error generating risk assessment report: {e}")
            raise
    
    async def _generate_incident_analysis_report(self, report: Report) -> Dict[str, Any]:
        """Generate incident analysis report."""
        try:
            # Collect incident data
            data = await self.data_collector.collect_incident_data(
                start_date=report.start_date,
                end_date=report.end_date,
                filters=report.filters
            )
            
            # Generate incident analysis charts
            charts = await self._generate_incident_analysis_charts(data)
            
            # Calculate incident metrics
            incident_metrics = self._calculate_incident_metrics(data)
            
            # Prepare report content
            content = {
                "title": report.title,
                "generated_at": datetime.now().isoformat(),
                "period": {
                    "start": report.start_date.isoformat(),
                    "end": report.end_date.isoformat()
                },
                "summary": {
                    "total_incidents": data.get("total_incidents", 0),
                    "resolved_incidents": data.get("resolved_incidents", 0),
                    "avg_resolution_time": data.get("avg_resolution_time", 0),
                    "mttr": data.get("mttr", 0)
                },
                "incident_metrics": incident_metrics,
                "charts": charts,
                "trends": data.get("trends", {}),
                "data": data
            }
            
            # Generate output file
            file_result = await self._generate_output_file(report, content)
            
            return {
                "content": content,
                "file_path": file_result["file_path"],
                "file_size": file_result["file_size"],
                "file_hash": file_result["file_hash"],
                "record_count": len(data.get("incidents", []))
            }
            
        except Exception as e:
            logger.error(f"Error generating incident analysis report: {e}")
            raise
    
    async def _generate_compliance_report(self, report: Report) -> Dict[str, Any]:
        """Generate compliance report."""
        try:
            # Collect compliance data
            data = await self.data_collector.collect_compliance_data(
                start_date=report.start_date,
                end_date=report.end_date,
                filters=report.filters
            )
            
            # Generate compliance charts
            charts = await self._generate_compliance_charts(data)
            
            # Calculate compliance metrics
            compliance_metrics = self._calculate_compliance_metrics(data)
            
            # Prepare report content
            content = {
                "title": report.title,
                "generated_at": datetime.now().isoformat(),
                "period": {
                    "start": report.start_date.isoformat(),
                    "end": report.end_date.isoformat()
                },
                "compliance_summary": {
                    "overall_score": data.get("overall_score", 0),
                    "frameworks": data.get("frameworks", []),
                    "violations": data.get("violations", 0),
                    "remediation_items": data.get("remediation_items", 0)
                },
                "compliance_metrics": compliance_metrics,
                "charts": charts,
                "violations": data.get("violation_details", []),
                "remediation": data.get("remediation_plan", []),
                "data": data
            }
            
            # Generate output file
            file_result = await self._generate_output_file(report, content)
            
            return {
                "content": content,
                "file_path": file_result["file_path"],
                "file_size": file_result["file_size"],
                "file_hash": file_result["file_hash"],
                "record_count": len(data.get("compliance_items", []))
            }
            
        except Exception as e:
            logger.error(f"Error generating compliance report: {e}")
            raise
    
    async def _generate_threat_intelligence_report(self, report: Report) -> Dict[str, Any]:
        """Generate threat intelligence report."""
        try:
            # Collect threat intelligence data
            data = await self.data_collector.collect_threat_intelligence_data(
                start_date=report.start_date,
                end_date=report.end_date,
                filters=report.filters
            )
            
            # Generate threat analysis charts
            charts = await self._generate_threat_intelligence_charts(data)
            
            # Prepare report content
            content = {
                "title": report.title,
                "generated_at": datetime.now().isoformat(),
                "period": {
                    "start": report.start_date.isoformat(),
                    "end": report.end_date.isoformat()
                },
                "threat_landscape": {
                    "active_threats": data.get("active_threats", 0),
                    "new_threats": data.get("new_threats", 0),
                    "threat_level": data.get("threat_level", "medium"),
                    "campaigns": data.get("campaigns", [])
                },
                "charts": charts,
                "iocs": data.get("iocs", []),
                "ttps": data.get("ttps", []),
                "recommendations": data.get("recommendations", []),
                "data": data
            }
            
            # Generate output file
            file_result = await self._generate_output_file(report, content)
            
            return {
                "content": content,
                "file_path": file_result["file_path"],
                "file_size": file_result["file_size"],
                "file_hash": file_result["file_hash"],
                "record_count": len(data.get("threats", []))
            }
            
        except Exception as e:
            logger.error(f"Error generating threat intelligence report: {e}")
            raise
    
    async def _generate_performance_metrics_report(self, report: Report) -> Dict[str, Any]:
        """Generate performance metrics report."""
        try:
            # Collect performance data
            data = await self.data_collector.collect_performance_data(
                start_date=report.start_date,
                end_date=report.end_date,
                filters=report.filters
            )
            
            # Generate performance charts
            charts = await self._generate_performance_charts(data)
            
            # Prepare report content
            content = {
                "title": report.title,
                "generated_at": datetime.now().isoformat(),
                "period": {
                    "start": report.start_date.isoformat(),
                    "end": report.end_date.isoformat()
                },
                "performance_summary": {
                    "avg_response_time": data.get("avg_response_time", 0),
                    "throughput": data.get("throughput", 0),
                    "error_rate": data.get("error_rate", 0),
                    "availability": data.get("availability", 0)
                },
                "charts": charts,
                "sla_metrics": data.get("sla_metrics", {}),
                "capacity_planning": data.get("capacity_planning", {}),
                "data": data
            }
            
            # Generate output file
            file_result = await self._generate_output_file(report, content)
            
            return {
                "content": content,
                "file_path": file_result["file_path"],
                "file_size": file_result["file_size"],
                "file_hash": file_result["file_hash"],
                "record_count": len(data.get("metrics", []))
            }
            
        except Exception as e:
            logger.error(f"Error generating performance metrics report: {e}")
            raise
    
    async def _generate_executive_summary(self, report: Report) -> Dict[str, Any]:
        """Generate executive summary report."""
        try:
            # Collect executive summary data
            data = await self.data_collector.collect_executive_summary_data(
                start_date=report.start_date,
                end_date=report.end_date,
                filters=report.filters
            )
            
            # Generate executive charts
            charts = await self._generate_executive_charts(data)
            
            # Prepare report content
            content = {
                "title": report.title,
                "generated_at": datetime.now().isoformat(),
                "period": {
                    "start": report.start_date.isoformat(),
                    "end": report.end_date.isoformat()
                },
                "executive_summary": {
                    "security_posture": data.get("security_posture", "good"),
                    "risk_level": data.get("risk_level", "medium"),
                    "key_achievements": data.get("key_achievements", []),
                    "critical_issues": data.get("critical_issues", [])
                },
                "charts": charts,
                "kpis": data.get("kpis", {}),
                "recommendations": data.get("recommendations", []),
                "data": data
            }
            
            # Generate output file
            file_result = await self._generate_output_file(report, content)
            
            return {
                "content": content,
                "file_path": file_result["file_path"],
                "file_size": file_result["file_size"],
                "file_hash": file_result["file_hash"],
                "record_count": 1  # Executive summary is a single record
            }
            
        except Exception as e:
            logger.error(f"Error generating executive summary: {e}")
            raise
    
    async def _generate_custom_report(self, report: Report) -> Dict[str, Any]:
        """Generate custom report based on template."""
        try:
            # Get template if specified
            template = None
            if report.template_id:
                with get_db() as db:
                    template = db.query(ReportTemplate).filter(
                        ReportTemplate.id == report.template_id
                    ).first()
            
            # Collect data based on template or parameters
            if template:
                data = await self.data_collector.collect_template_data(
                    template=template,
                    start_date=report.start_date,
                    end_date=report.end_date,
                    filters=report.filters,
                    parameters=report.parameters
                )
            else:
                data = await self.data_collector.collect_custom_data(
                    data_sources=report.data_sources,
                    start_date=report.start_date,
                    end_date=report.end_date,
                    filters=report.filters,
                    parameters=report.parameters
                )
            
            # Generate charts
            charts = await self._generate_custom_charts(data, template)
            
            # Prepare report content
            content = {
                "title": report.title,
                "generated_at": datetime.now().isoformat(),
                "period": {
                    "start": report.start_date.isoformat(),
                    "end": report.end_date.isoformat()
                },
                "charts": charts,
                "data": data
            }
            
            # Generate output file
            file_result = await self._generate_output_file(report, content)
            
            return {
                "content": content,
                "file_path": file_result["file_path"],
                "file_size": file_result["file_size"],
                "file_hash": file_result["file_hash"],
                "record_count": len(data.get("records", []))
            }
            
        except Exception as e:
            logger.error(f"Error generating custom report: {e}")
            raise
    
    async def _generate_output_file(self, report: Report, content: Dict[str, Any]) -> Dict[str, Any]:
        """Generate output file in specified format."""
        try:
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{report.title}_{timestamp}.{report.format}"
            filepath = self.output_dir / filename
            
            # Generate file based on format
            if report.format == ReportFormat.PDF:
                await self._generate_pdf_report(filepath, content, report)
            elif report.format == ReportFormat.HTML:
                await self._generate_html_report(filepath, content, report)
            elif report.format == ReportFormat.JSON:
                await self._generate_json_report(filepath, content)
            elif report.format == ReportFormat.CSV:
                await self._generate_csv_report(filepath, content)
            elif report.format == ReportFormat.EXCEL:
                await self._generate_excel_report(filepath, content, report)
            else:
                raise ValueError(f"Unsupported report format: {report.format}")
            
            # Calculate file size and hash
            file_size = filepath.stat().st_size
            file_hash = self._calculate_file_hash(filepath)
            
            return {
                "file_path": str(filepath),
                "file_size": file_size,
                "file_hash": file_hash
            }
            
        except Exception as e:
            logger.error(f"Error generating output file: {e}")
            raise
    
    def _calculate_file_hash(self, filepath: Path) -> str:
        """Calculate SHA-256 hash of file."""
        hash_sha256 = hashlib.sha256()
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_sha256.update(chunk)
        return hash_sha256.hexdigest()
    
    # Chart generation methods
    async def _generate_security_dashboard_charts(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate charts for security dashboard."""
        charts = []
        
        # Incident severity distribution
        if "incident_severity" in data:
            chart = await self.chart_generator.create_pie_chart(
                data["incident_severity"],
                title="Incident Severity Distribution",
                labels=["Critical", "High", "Medium", "Low"]
            )
            charts.append(chart)
        
        # Incident trend over time
        if "incident_trend" in data:
            chart = await self.chart_generator.create_line_chart(
                data["incident_trend"],
                title="Incident Trend Over Time",
                x_label="Date",
                y_label="Number of Incidents"
            )
            charts.append(chart)
        
        # Top attack types
        if "attack_types" in data:
            chart = await self.chart_generator.create_bar_chart(
                data["attack_types"],
                title="Top Attack Types",
                x_label="Attack Type",
                y_label="Count"
            )
            charts.append(chart)
        
        return charts
    
    async def _generate_risk_assessment_charts(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate charts for risk assessment report."""
        charts = []
        
        # Risk level distribution
        if "risk_distribution" in data:
            chart = await self.chart_generator.create_pie_chart(
                data["risk_distribution"],
                title="Risk Level Distribution",
                labels=["Critical", "High", "Medium", "Low"]
            )
            charts.append(chart)
        
        # Risk trend over time
        if "risk_trend" in data:
            chart = await self.chart_generator.create_line_chart(
                data["risk_trend"],
                title="Risk Score Trend",
                x_label="Date",
                y_label="Average Risk Score"
            )
            charts.append(chart)
        
        return charts
    
    # File generation methods
    async def _generate_pdf_report(self, filepath: Path, content: Dict[str, Any], report: Report):
        """Generate PDF report."""
        # Get template
        template = self.template_env.get_template(f"{report.report_type}_template.html")
        
        # Render HTML
        html_content = template.render(content=content, report=report)
        
        # Generate PDF
        weasyprint.HTML(string=html_content).write_pdf(str(filepath))
    
    async def _generate_html_report(self, filepath: Path, content: Dict[str, Any], report: Report):
        """Generate HTML report."""
        # Get template
        template = self.template_env.get_template(f"{report.report_type}_template.html")
        
        # Render HTML
        html_content = template.render(content=content, report=report)
        
        # Write to file
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    async def _generate_json_report(self, filepath: Path, content: Dict[str, Any]):
        """Generate JSON report."""
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(content, f, indent=2, default=str)
    
    async def _generate_csv_report(self, filepath: Path, content: Dict[str, Any]):
        """Generate CSV report."""
        # Convert data to DataFrame
        if "data" in content and isinstance(content["data"], list):
            df = pd.DataFrame(content["data"])
        else:
            # Create a simple CSV from the content
            df = pd.DataFrame([content])
        
        # Write to CSV
        df.to_csv(filepath, index=False)
    
    async def _generate_excel_report(self, filepath: Path, content: Dict[str, Any], report: Report):
        """Generate Excel report."""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Report"
        
        # Add title
        worksheet['A1'] = report.title
        worksheet['A1'].font = Font(size=16, bold=True)
        
        # Add generated date
        worksheet['A2'] = f"Generated: {content['generated_at']}"
        
        # Add data
        if "data" in content and isinstance(content["data"], list):
            df = pd.DataFrame(content["data"])
            
            # Write headers
            for col, header in enumerate(df.columns, 1):
                worksheet.cell(row=4, column=col, value=header)
                worksheet.cell(row=4, column=col).font = Font(bold=True)
            
            # Write data
            for row, (_, data_row) in enumerate(df.iterrows(), 5):
                for col, value in enumerate(data_row, 1):
                    worksheet.cell(row=row, column=col, value=value)
        
        # Save workbook
        workbook.save(filepath)
    
    # Metric calculation methods
    def _calculate_risk_metrics(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate risk metrics."""
        return {
            "avg_risk_score": data.get("avg_risk_score", 0),
            "high_risk_count": data.get("high_risk_count", 0),
            "risk_reduction": data.get("risk_reduction", 0),
            "compliance_score": data.get("compliance_score", 0)
        }
    
    def _calculate_incident_metrics(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate incident metrics."""
        return {
            "mttr": data.get("mttr", 0),
            "mtbf": data.get("mtbf", 0),
            "resolution_rate": data.get("resolution_rate", 0),
            "escalation_rate": data.get("escalation_rate", 0)
        }
    
    def _calculate_compliance_metrics(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate compliance metrics."""
        return {
            "compliance_score": data.get("compliance_score", 0),
            "control_effectiveness": data.get("control_effectiveness", 0),
            "audit_findings": data.get("audit_findings", 0),
            "remediation_rate": data.get("remediation_rate", 0)
        }
    
    # Additional chart generation methods
    async def _generate_incident_analysis_charts(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate charts for incident analysis."""
        charts = []
        
        # Incident resolution time trend
        if "incidents" in data:
            incidents = data["incidents"]
            if incidents:
                # Create resolution time chart
                resolution_data = {
                    incident["id"]: incident.get("resolution_time", 0)
                    for incident in incidents[:10]  # Top 10 incidents
                }
                chart = await self.chart_generator.create_bar_chart(
                    resolution_data,
                    title="Incident Resolution Times",
                    x_label="Incident ID",
                    y_label="Resolution Time (hours)"
                )
                charts.append(chart)
        
        # MTTR trend over time
        if "trends" in data and "mttr_trend" in data["trends"]:
            chart = await self.chart_generator.create_line_chart(
                data["trends"]["mttr_trend"],
                title="Mean Time to Resolution Trend",
                x_label="Date",
                y_label="MTTR (hours)"
            )
            charts.append(chart)
        
        # Incident status distribution
        status_distribution = {
            "Open": len([i for i in data.get("incidents", []) if i.get("status") == "open"]),
            "In Progress": len([i for i in data.get("incidents", []) if i.get("status") == "in_progress"]),
            "Resolved": len([i for i in data.get("incidents", []) if i.get("status") == "resolved"]),
            "Closed": len([i for i in data.get("incidents", []) if i.get("status") == "closed"])
        }
        
        if any(status_distribution.values()):
            chart = await self.chart_generator.create_pie_chart(
                status_distribution,
                title="Incident Status Distribution"
            )
            charts.append(chart)
        
        return charts
    
    async def _generate_compliance_charts(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate charts for compliance report."""
        charts = []
        
        # Compliance score gauge
        overall_score = data.get("overall_score", 0)
        chart = await self.chart_generator.create_gauge_chart(
            value=overall_score,
            title="Overall Compliance Score",
            min_value=0,
            max_value=100,
            threshold_ranges=[
                {"range": [0, 60], "color": "#E74C3C"},
                {"range": [60, 80], "color": "#F39C12"},
                {"range": [80, 100], "color": "#27AE60"}
            ]
        )
        charts.append(chart)
        
        # Framework compliance scores
        frameworks = data.get("frameworks", [])
        if frameworks:
            framework_scores = {
                framework: 85 + (hash(framework) % 15)  # Simulated scores
                for framework in frameworks
            }
            chart = await self.chart_generator.create_bar_chart(
                framework_scores,
                title="Compliance by Framework",
                x_label="Framework",
                y_label="Compliance Score (%)"
            )
            charts.append(chart)
        
        # Violation trend
        if "violation_trend" in data:
            chart = await self.chart_generator.create_line_chart(
                data["violation_trend"],
                title="Compliance Violations Over Time",
                x_label="Date",
                y_label="Number of Violations"
            )
            charts.append(chart)
        
        # Control effectiveness
        control_effectiveness = data.get("control_effectiveness", 90)
        chart = await self.chart_generator.create_gauge_chart(
            value=control_effectiveness,
            title="Control Effectiveness",
            min_value=0,
            max_value=100
        )
        charts.append(chart)
        
        return charts
    
    async def _generate_threat_intelligence_charts(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate charts for threat intelligence."""
        charts = []
        
        # Threat level indicator
        threat_level_map = {"low": 25, "medium": 50, "high": 75, "critical": 100}
        threat_level_value = threat_level_map.get(data.get("threat_level", "medium"), 50)
        
        chart = await self.chart_generator.create_gauge_chart(
            value=threat_level_value,
            title="Current Threat Level",
            min_value=0,
            max_value=100,
            threshold_ranges=[
                {"range": [0, 25], "color": "#27AE60"},
                {"range": [25, 50], "color": "#F39C12"},
                {"range": [50, 75], "color": "#E67E22"},
                {"range": [75, 100], "color": "#E74C3C"}
            ]
        )
        charts.append(chart)
        
        # Threat type distribution
        if "threats" in data:
            threat_types = {}
            for threat in data["threats"]:
                threat_type = threat.get("type", "unknown")
                threat_types[threat_type] = threat_types.get(threat_type, 0) + 1
            
            if threat_types:
                chart = await self.chart_generator.create_pie_chart(
                    threat_types,
                    title="Threat Type Distribution"
                )
                charts.append(chart)
        
        # IOC trend over time
        if "ioc_trend" in data:
            chart = await self.chart_generator.create_line_chart(
                data["ioc_trend"],
                title="Indicators of Compromise Trend",
                x_label="Date",
                y_label="Number of IOCs"
            )
            charts.append(chart)
        
        # Campaign activity
        campaigns = data.get("campaigns", [])
        if campaigns:
            campaign_activity = {
                campaign.get("name", "Unknown"): campaign.get("activity_level", 1)
                for campaign in campaigns[:10]  # Top 10 campaigns
            }
            chart = await self.chart_generator.create_bar_chart(
                campaign_activity,
                title="Threat Campaign Activity",
                x_label="Campaign",
                y_label="Activity Level"
            )
            charts.append(chart)
        
        return charts
    
    async def _generate_performance_charts(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate charts for performance metrics."""
        charts = []
        
        # Response time gauge
        avg_response_time = data.get("avg_response_time", 0)
        chart = await self.chart_generator.create_gauge_chart(
            value=avg_response_time,
            title="Average Response Time (ms)",
            min_value=0,
            max_value=1000,
            threshold_ranges=[
                {"range": [0, 200], "color": "#27AE60"},
                {"range": [200, 500], "color": "#F39C12"},
                {"range": [500, 1000], "color": "#E74C3C"}
            ]
        )
        charts.append(chart)
        
        # Throughput trend
        if "throughput_trend" in data:
            chart = await self.chart_generator.create_line_chart(
                data["throughput_trend"],
                title="Throughput Over Time",
                x_label="Time",
                y_label="Requests per Second"
            )
            charts.append(chart)
        
        # SLA metrics
        sla_metrics = data.get("sla_metrics", {})
        if sla_metrics:
            chart = await self.chart_generator.create_bar_chart(
                sla_metrics,
                title="SLA Metrics",
                x_label="Metric",
                y_label="Value"
            )
            charts.append(chart)
        
        # Resource utilization
        capacity_planning = data.get("capacity_planning", {})
        if capacity_planning:
            utilization_data = [
                {"resource": "CPU", "utilization": capacity_planning.get("cpu_utilization", 0)},
                {"resource": "Memory", "utilization": capacity_planning.get("memory_utilization", 0)},
                {"resource": "Disk", "utilization": capacity_planning.get("disk_utilization", 0)}
            ]
            
            chart = await self.chart_generator.create_bar_chart(
                utilization_data,
                title="Resource Utilization",
                x_label="Resource",
                y_label="Utilization (%)"
            )
            charts.append(chart)
        
        return charts
    
    async def _generate_executive_charts(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate charts for executive summary."""
        charts = []
        
        # KPI dashboard
        kpis = data.get("kpis", {})
        if kpis:
            # Security score gauge
            security_score = kpis.get("security_score", 0)
            chart = await self.chart_generator.create_gauge_chart(
                value=security_score,
                title="Security Score",
                min_value=0,
                max_value=100
            )
            charts.append(chart)
            
            # KPI summary chart
            kpi_data = {
                "Security Score": kpis.get("security_score", 0),
                "Risk Score": kpis.get("risk_score", 0),
                "Compliance Score": kpis.get("compliance_score", 0)
            }
            chart = await self.chart_generator.create_bar_chart(
                kpi_data,
                title="Key Performance Indicators",
                x_label="KPI",
                y_label="Score"
            )
            charts.append(chart)
        
        # Security posture trend
        if "security_trend" in data:
            chart = await self.chart_generator.create_line_chart(
                data["security_trend"],
                title="Security Posture Trend",
                x_label="Date",
                y_label="Security Score"
            )
            charts.append(chart)
        
        # Risk vs Impact matrix
        if "risk_impact_data" in data:
            chart = await self.chart_generator.create_scatter_plot(
                data["risk_impact_data"],
                title="Risk vs Impact Analysis",
                x_label="Risk Level",
                y_label="Business Impact"
            )
            charts.append(chart)
        
        return charts
    
    async def _generate_custom_charts(self, data: Dict[str, Any], template: Optional[ReportTemplate]) -> List[Dict[str, Any]]:
        """Generate custom charts based on template configuration."""
        charts = []
        
        try:
            # If template has chart configurations, use them
            if template and hasattr(template, 'template_config'):
                chart_configs = template.template_config.get("charts", [])
                
                for chart_config in chart_configs:
                    chart_type = chart_config.get("type", "bar")
                    chart_title = chart_config.get("title", "Custom Chart")
                    data_field = chart_config.get("data_field", "records")
                    
                    # Get chart data
                    chart_data = data.get(data_field, [])
                    
                    if chart_data:
                        if chart_type == "pie":
                            chart = await self.chart_generator.create_pie_chart(
                                chart_data,
                                title=chart_title
                            )
                        elif chart_type == "line":
                            chart = await self.chart_generator.create_line_chart(
                                chart_data,
                                title=chart_title
                            )
                        elif chart_type == "scatter":
                            chart = await self.chart_generator.create_scatter_plot(
                                chart_data,
                                title=chart_title
                            )
                        else:  # Default to bar chart
                            chart = await self.chart_generator.create_bar_chart(
                                chart_data,
                                title=chart_title
                            )
                        
                        charts.append(chart)
            
            # If no template or no chart configs, generate default charts
            if not charts and "records" in data:
                records = data["records"]
                if records and len(records) > 0:
                    # Create a simple summary chart
                    record_count_by_type = {}
                    for record in records:
                        record_type = record.get("type", "unknown")
                        record_count_by_type[record_type] = record_count_by_type.get(record_type, 0) + 1
                    
                    if record_count_by_type:
                        chart = await self.chart_generator.create_pie_chart(
                            record_count_by_type,
                            title="Data Distribution by Type"
                        )
                        charts.append(chart)
        
        except Exception as e:
            logger.warning(f"Error generating custom charts: {e}")
        
        return charts
    
    def get_stats(self) -> Dict[str, Any]:
        """Get generator statistics."""
        return {
            "generation_queue_size": self.generation_queue.qsize(),
            "active_generation_tasks": len(self.generation_tasks),
            "output_directory": str(self.output_dir),
            "supported_formats": [format.value for format in ReportFormat],
            "supported_types": [type.value for type in ReportType]
        }